# -*- coding: utf-8 -*-
"""
Created on Fri Jun 11 15:43:59 2021

@author: chaibou001
"""

#https://www.marsja.se/your-guide-to-reading-excel-xlsx-files-in-python/

from pylab import*
from matplotlib import*
from openpyxl import* #modules import
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from matplotlib.pyplot import*
import datetime

#path='C:\\Users\\chaibou001\\Desktop\\Rotation R&D\\Raw_data'

#savep='C:\\Users\\chaibou001\\Desktop\\Rotation R&D\\Figures\\'

clean_columns=[
'N Test Enac',
'Code CA Manufacturer Plant',
'Production order (PO)',
'Drum',
'Cable',
'Diameter',
'Number of cables',
'Laboratory',
'Compound',
'Open or Closed Doors Humidity content',
'Humidity',
'Temperature',
'Safety Margin',
'Test date',
'Euroclasse',
'FS',
'THR1200s',
'HRRmax',
'Time',
'FIGRA',
'Euroclass',
'TSP1200s',
'Peak SPR',
'Smoke',
'Flaming inf 10s',
'Flaming sup 10s',
'Droplets']


################## Scripts in Pandas #####################
'''
def read_table_pandas_base(p,title,sheet_name):
    print('Reading file...')
    tic=time.time()
    # read excel sheet, skips first 7 lines
    df = pd.read_excel(p+'\\'+title+'.xlsx',sheet_name,skiprows=6)
    # clean columns
    df.columns=clean_columns
    elapsed= time.time()-tic
    print(f'- Successful. Time elapsed: {elapsed:2.3} seconds\n')
    return(df)

def read_table_pandas_clean(p,title,sheet_name):
    print('Reading file...')
    tic=time.time()
    # read excel sheet
    df = pd.read_excel(p+'\\'+title+'.xlsx',sheet_name)
    elapsed= time.time()-tic
    print(f'- Successful. Time elapsed: {elapsed:2.3} seconds\n')
    return(df)
'''


def read_table_dirty(path):
    print('Reading file...')
    tic=time.time()
    # read excel sheet, skips first 7 lines
    df = pd.read_excel(path,0,skiprows=6)
    # clean columns
    df.columns=clean_columns
    elapsed= time.time()-tic
    print(f'- Successful. Time elapsed: {elapsed:2.3} seconds\n')
    return(df)


def read_table(path):
    print('Reading file...')
    tic=time.time()
    # read excel sheet
    df = pd.read_excel(path,0)
    elapsed= time.time()-tic
    print(f'- Successful. Time elapsed: {elapsed:2.3} seconds\n')
    return(df)

def clean_table(df,path,save_file):
    print('Cleaning file...')
    tic=time.time()
    # create a new dataframe clean of lines with no cable name or humidity
    # deletes: (1) repeated lines due to merging (2) lines with no useful data (3) last empty block of lines
    df_clean=df.loc[(pd.isnull(df['Cable'])==False) & (pd.isnull(df['Humidity'])==False)  & (pd.isnull(df['Temperature'])==False) ]

    df_clean=df_clean[df_clean['Temperature']!='--']
    df_clean=df_clean[df_clean['Temperature']!='??']
    df_clean=df_clean[df_clean['Temperature']!='NC']   

    df_clean=df_clean[df_clean['Humidity']!='--']
    df_clean=df_clean[df_clean['Humidity']!='??']
    df_clean=df_clean[df_clean['Humidity']!='NC']
    '''
    df_clean['Temperature']=df_clean['Temperature'].str.replace(',','.')
    df_clean['Humidity']=df_clean['Humidity'].str.replace(',','.')
    '''
    
    df_clean=df_clean[df_clean['FS']!='-']
    df_clean=df_clean[df_clean['FS']!='--']
    df_clean=df_clean[df_clean['FS']!='??']
    df_clean=df_clean[df_clean['THR1200s']!='-']    
    df_clean=df_clean[df_clean['THR1200s']!='--']
    df_clean=df_clean[df_clean['THR1200s']!='??']

    df_clean=df_clean[df_clean['FIGRA']!='-']
    df_clean=df_clean[df_clean['FIGRA']!='--']
    df_clean=df_clean[df_clean['FIGRA']!='??']
    df_clean=df_clean[df_clean['FS']!='-']    
    df_clean=df_clean[df_clean['HRRmax']!='--']
    df_clean=df_clean[df_clean['HRRmax']!='??']
    
    # remove spaces (x here is a column, we can add index=1 if we want to navigate through rows)    
    df_clean['Cable']=df_clean['Cable'].str.strip().str.upper().str.replace(',','.').str.replace(' AS ',' (AS) ').str.replace('1000V','1KV').str.replace('1000 V','1KV')
    df_clean['Compound']=df_clean['Compound'].str.replace(' ','').str.replace('(','').str.replace(')','').str.replace(',','.').str.upper()

    '''
    df_clean['Cable']=df_clean['Cable'].apply(
    lambda x: x.str.strip() if type(x)==string
    )
    df_clean['Laboratory']=df_clean['Laboratory'].apply(
    lambda x: x.str.strip() if type(x)==string
    )
    df_clean['Compound']=df_clean['Compound'].apply(
    lambda x: x.str.strip() if type(x)==string
    )
    '''    
    # create excel writer
    path='C:/Users/chaibou001/Desktop'
    writer = pd.ExcelWriter(path+'/'+save_file+'.xlsx')
    # write dataframe to excel sheet named save_file
    df_clean.to_excel(writer,index=False)
    # save the excel file
    writer.save()
    elapsed= time.time()-tic
    print(f'- Successful. Time elapsed: {elapsed:2.3} seconds\n')    
    print('Saved successfully to Excel sheet: '+save_file+'.xlsx')
    return 0

def statistics(df):
    print('[*] Statistics\n')
    '''
    site_distribution=df['Laboratory'].value_counts()
    cable_distribution=df['Cable'].str.upper().value_counts()
    cmp_distribution=(df['Compound'].str.replace(' ','').str.upper()).value_counts()
    '''
    site_distribution=df['Laboratory'].value_counts()
    cable_distribution=df['Cable'].value_counts()
    cmp_distribution=df['Compound'].value_counts()
    a=len(df)
    b=site_distribution
    c=cable_distribution[0:20]#.to_string()
    d=cmp_distribution[0:20]#.to_string()
    print('Fire test count:',a)
    print('\n\n')
    print('Site distribution:\n') 
    print(b) # Insert Pie chart here for data vizualization: will do it later
    print('\n\n')
    print('Cable distribution:\n')
    print(c)
    print('\n\n')
    print('Compound distribution:\n')
    print(d)
    print('\n\n')
    return a,b,c,d

def search(df,site,cable,compound):
    # another way to do it: look at the name of variables and try to play on their formatting
    a={site: 'Laboratory', cable: 'Cable', compound: 'Compound'}
    df_target=df
    for j,k in enumerate([site, cable, compound]):
        if(k!=''):
            df_target=df_target[df_target[a[k]]==k]
    return(df_target)

#r"""
def plot_data(df,site,cable,compound,y,x,l):
    #df=read_table(r'C:\Users\chaibou001\Desktop','Relations essais_clean')
    a={site: 'Laboratory', cable: 'Cable', compound: 'Compound'}
    df_target=df
    # Filter based on site + cable + compound
    for j,k in enumerate([site, cable, compound]):
        if(k!=''):
            df_target=df_target[df_target[a[k]]==k]
    # Filter based on x,y,l
    if(y=='HRRpeak'):
        y='HRRmax'
    y_axis=df_target[y]
    x_axis=df_target[x]
    if l!='nolegend':
        l_arr=df_target[l]
    else:
        l_arr=0
    return x_axis, y_axis, l_arr


################## Scripts in Openpyxl #####################

# Function: Reads an Excel (.xlsx) file and ports it into your code
# Input: Title of the Excel file containing the data (has to be in the path)
# Output: List that contains Excel sheet data   
'''
def read_table(title):
    tic=time.time()
    file_name='\\'+title+'.xlsx'    
    book=load_workbook(path+file_name,data_only=True)
    sheet=book.active
   
    data=[[] for _ in range(sheet.max_column)]
    i=0
    for column in sheet.iter_cols(min_row=8): 
        for j in range(0,sheet.max_row-7):
            data[i].append(column[j].value)
        i=(i+1)      
    elapsed= time.time()-tic
    print(f'- Time elapsed: {elapsed:2.2} seconds\n')
    return(data)

# Function: Skips headline, empty lines and saves cleaned Excel data
# Input: List containing Excel sheet data, Save file name
# Output: None

def clean_it(data_in,cleaned_it):
    tic=time.time()    
    datax=[[] for _ in range(len(data_in))]
    
    for j in range(0,len(datax)):
        for i in range(1,len(data_in[0])):
            if(data_in[4][i] is not None):
                datax[j].append(data_in[j][i].strip())
                
    wb=Workbook()
    ws1=wb[wb.sheetnames[0]]    
    for j in range(1,len(datax)+1):
        for i in range(1,len(datax[0])+1):
            ws1.cell(i,j).value=datax[j-1][i-1]    
    wb.save(path+'\\'+cleaned_it+'.xlsx')    
    elapsed= time.time()-tic
    print(f'- Time elapsed: {elapsed:.2} seconds\n')
    
def adjust_table(data):
    stop=[i for i,x in enumerate(data[15]) if data[15][i]!=' ' and data[15][i-1]!=' '][-1]
    data2=data
    for j in range(0,len(data)):
        data2[j]=data[j][0:stop+1]
                
    for j in range(0,len(data)):
        for i in range(1,len(data[0])):
            if(data[j][i] is None):
                data2[j][i]=data[j][i-1]    
                
    wb=Workbook()
    ws1=wb[wb.sheetnames[0]]    
    for j in range(1,len(data2)+1):
        for i in range(1,stop+1):
            ws1.cell(i,j).value=data2[j-1][i-1]    
    wb.save(path+'\\Book1.xlsx')    
    return data2

def stats(data3):
    test_count=len(data3[0])
    # Logging unique locations/sites
    unique_sites=list(set(data3[7]))
    unique_locations=[[] for _ in range(len(unique_sites))]
    for i in range(0,len(unique_sites)):
        # Transposing the location list
        unique_locations[i].append(unique_sites[i])
        # Indices of all occurences of the unique locations in the Excel file
        indices = [j for j, x in enumerate(data3[7]) if x == unique_locations[i][0]]
        # Test count of each unique location
        unique_locations[i].append(len(indices)) 
        # Number of unique cables in each unique location
        unique_locations[i].append(len(list(set([e for i, e in enumerate(data3[4]) if i in indices]))))
    # Code of the unique cables    
    #unique_cables=set(data3[4])
    print(f'[*] {test_count} fire tests between {data3[13][1].date()} and {data3[13][-1].date()}')
    #print(f'[*] {len(unique_cables)} unique cable(s)')   
    print(f'[*] {len(unique_locations)} unique location(s):')
    for i in range(0,len(unique_sites)):
            print(f'    {unique_locations[i][0]}, {unique_locations[i][1]} fire test(s)')  
    return(unique_locations)

def tests(data,loc):
# https://realpython.com/python-enumerate/
# j,x in enumerate(index,val) output   
    indices=[j for j, x in enumerate(data[7]) if x == loc]   
    cables=([data[4][i] for i in indices])
    unique_cables=list(set(cables))
    index=[[] for _ in range(len(unique_cables))]
    cnt=zeros(len(unique_cables))
    for i, j in enumerate(unique_cables): # enumerates the elements, i is a str not an int counter
        for k, l in enumerate(indices):
            if(data[4][l]==j):
                cnt[i]=cnt[i]+1;

    sorted_args=argsort(cnt)[::-1]
    print('\n[*] Most tested cable(s): \n')
    most_tested=[]
    for i,j in enumerate(sorted_args[0:30]):
        most_tested.append(unique_cables[j])
        print(f'[{i}] Cable "{unique_cables[j]}", {int(cnt[j])} fire tests.')                            
    return(most_tested)

def compounds(data,loc,cable_name):
# https://realpython.com/python-enumerate/
# j,x in enumerate(index,val) output   
    indices=[j for j, x in enumerate(data[7]) if x == loc and data[4][j]==cable_name]
    cmp=[data[8][i] for i in indices]
    cmp=list(set(cmp))
    index=[[] for _ in range(len(cmp))]
    cnt=zeros(len(cmp))
    for i, j in enumerate(cmp): # enumerates the elements, i is a str not an int counter
        for k, l in enumerate(indices):
            if(data[8][l]==j):
                cnt[i]=cnt[i]+1;
    sorted_args=argsort(cnt)[::-1]             
    print(f'\n[*] Cable: {cable_name} \n[*] Compound(s): \n')
    for i,j in enumerate(sorted_args[:]):
        print(f'[{i}] Compound "{cmp[j]}", {int(cnt[j])} fire test(s).')
        cmp[i]=cmp[j]
    return(cmp)
    
def extract_data(data,loc,cable_name,cmp):
    indices=[j for j, x in enumerate(data[7]) if x == loc and data[4][j]==cable_name and data[8][j]==cmp]
    extracted_data=[[] for _ in range(len(data))]
    for i in range(len(data)):
        extracted_data[i].append([data[i][k] for k in indices])
    return(extracted_data)

def save_data(site,folder):
    savep2=savep+str(site)+'\\'
    save_path=os.path.join(savep2,folder)
    os.makedirs(save_path,exist_ok=True)
    return(save_path)

def single(loc,cab,cmp,data3):
    spaces = re.compile(r'\s+')
    cmp=spaces.sub('',cmp)
    cab=spaces.sub('',cab)
    indices=[j for j, x in enumerate(data3[4]) if data3[7][j].strip().lower()==loc.strip().lower() and spaces.sub('',x).lower()==cab.lower() and spaces.sub('',data3[8][j]).lower()==cmp.lower() and data3[10][j] is not None]
    dates=[data3[13][x].date() for j,x in enumerate(indices)]
    return(dates,indices)

def read_table2(title):
    tic = time.time()
    file_name='\\'+title+'.xlsx'    
    book=load_workbook(path+file_name,data_only=True)
    sheet=book.active
    
    col_names = []
    
    # index in column[index] is the row number
    # for loop allows you to navigate through the columns through iter_cols = called an iterative generator
    # for more doc on iter generators: https://wiki.python.org/moin/Generators
    
    #iter_rows makes i print rows, iter_cols vice versa    
    #print(column) # is a tuple variable kinda like a list where the col elements can be accessed through i[n]  
    
    # is the loop here really necessary?
    for column in sheet.iter_cols(): # navigate through cols
        col_names.append(column[6].value)               # navigate through 6th row

    
    data=[[] for _ in range(sheet.max_column)]
    i=0
    for column in sheet.iter_cols(): 
        for j in range(0,sheet.max_row):
            data[i].append(column[j].value)
        i=(i+1)    
    
    print('Data loaded successfully')
    elapsed= time.time()-tic
    print(f'- Time elapsed: {elapsed:.3} seconds\n')
    return(data,col_names)
'''

